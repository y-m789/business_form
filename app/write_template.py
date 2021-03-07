import math
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as excImage


class WriteTemplate:
    def __init__(self, num_item):
        self.estimate_path = './app/xlsx_template/template_estimate.xlsx'
        self.bill_path = './app/xlsx_template/template_bill.xlsx'
        self.delivery_path = './app/xlsx_template/template_delivery.xlsx'
        self.num_item = num_item
        self.cell_init = 18  # 品目のセル番号

    def replace_cell(self, ws, base, cont):
        """
        テンプレートの対応するセルを入力された情報に置き換える
        :param ws:
            xlsxのシート
        :param base:
            入力された基本情報
        :param cont:
            入力された取引先情報
        :return:
            セルが置き換えられたxlsxのシート
        """
        total_price = 0
        total_tax = 0

        # 基本情報
        ws['K6'].value = base.company  # 会社名
        ws['K7'].value = base.post     # 郵便番号
        ws['K8'].value = base.address  # 住所
        ws['M10'].value = base.tel     # 電話番号
        ws['M11'].value = base.email   # E-mail
        ws['M12'].value = base.name    # 担当者名

        # 取引先情報
        ws['A3'].value = cont.customer+cont.customer_department  # 取引先会社名+部署名
        ws['D4'].value = cont.customer_name  # 担当者名
        ws['C6'].value = cont.title          # 件名
        for i in range(self.num_item):
            no_cell = 'A'+str(self.cell_init+i)
            item_cell = 'B'+str(self.cell_init+i)
            volume_cell = 'J'+str(self.cell_init+i)
            unit_cell = 'K'+str(self.cell_init+i)
            price_cell = 'L'+str(self.cell_init+i)
            total_cell = 'O'+str(self.cell_init+i)
            item_name = 'item'+str(i+1)
            volume_name = 'volume'+str(i+1)
            unit_name = 'unit'+str(i+1)
            price_name = 'price'+str(i+1)
            tax_name = 'tax'+str(i+1)

            # 数量・金額が入力されていれば置き換える
            if not cont.__dict__[volume_name] == '' and not cont.__dict__[price_name] == '':
                ws[item_cell].value = cont.__dict__[item_name]      # 品目
                ws[volume_cell].value = cont.__dict__[volume_name]  # 数量
                ws[unit_cell].value = cont.__dict__[unit_name]      # 単位
                ws[price_cell].value = cont.__dict__[price_name]    # 単価
                ws[no_cell].value = i+1                             # No.
                ws[total_cell].value = int(cont.__dict__[volume_name]) * int(cont.__dict__[price_name])  # 金額
                total_price += ws[total_cell].value       # 小計の計算
                total_tax += math.floor(ws[total_cell].value * float(cont.__dict__[tax_name])/100)
        ws['L30'].value = int(total_price)                   # 小計
        ws['L31'].value = int(total_tax)                     # 消費税
        ws['L32'].value = ws['L30'].value + ws['L31'].value  # 合計
        ws['D15'].value = ws['L32'].value                    # 合計金額

        return ws

    def estimate(self, base, cont):
        """
        見積書が選択された場合の処理
        見積書のテンプレートを読み込み、
        共通部分をload_workbook関数で置き換え、
        見積書専用部分をここで置き換える
        :param base:
            入力された基本情報
        :param cont:
            入力された取引先情報
        """
        # template読み込み
        wb = load_workbook(self.estimate_path)
        ws = self.replace_cell(wb['Sheet1'], base, cont)

        # 見積書のみの部分
        ws['N3'].value = cont.num_estimate        # 見積書番号
        ws['N4'].value = cont.timestamp_estimate  # 見積日
        ws['C36'].value = base.memo_estimate      # 備考

        # 書き込みしたファイルの一時保存
        wb.save('./tmp.xlsx')

    def bill(self, base, cont):
        """
        請求書が選択された場合の処理
        請求書のテンプレートを読み込み、
        共通部分をload_workbook関数で置き換え、
        請求書専用部分をここで置き換える
        :param base:
            入力された基本情報
        :param cont:
            入力された取引先情報
        """
        # template読み込み
        wb = load_workbook(self.bill_path)
        ws = self.replace_cell(wb['Sheet1'], base, cont)

        # 請求書のみの部分
        ws['B32'].value = base.bank           # 振込先
        ws['N3'].value = cont.num_bill        # 請求書番号
        ws['N4'].value = cont.timestamp_bill  # 請求日
        ws['M15'].value = cont.limit_bill     # 支払期限
        ws['C36'].value = base.memo_bill      # 備考

        # 書き込みしたファイルの一時保存
        wb.save('./tmp.xlsx')

    def delivery(self, base, cont):
        """
        納品書が選択された場合の処理
        納品書のテンプレートを読み込み、
        共通部分をload_workbook関数で置き換え、
        納品書専用部分をここで置き換える
        :param base:
            入力された基本情報
        :param cont:
            入力された取引先情報
        """
        # template読み込み
        wb = load_workbook(self.delivery_path)
        ws = self.replace_cell(wb['Sheet1'], base, cont)

        # 納品書のみの部分
        ws['N3'].value = cont.num_deliver        # 納品書番号
        ws['N4'].value = cont.timestamp_deliver  # 納品日
        ws['C36'].value = base.memo_deliver      # 備考

        # 書き込みしたファイルの一時保存
        wb.save('./tmp.xlsx')
